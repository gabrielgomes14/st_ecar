import matplotlib.pyplot as plt
import pandas as pd
import warnings
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils.dataframe import dataframe_to_rows
import streamlit as st
from io import BytesIO

warnings.filterwarnings("ignore")

def importar_planilha(file):
    OUTPUT_PATH = "relatorio_divergencias.xlsx"
    
    COLUMNS = [
        "Colaborador", "Executores",
        "Modelo de checklist", "Veículo",
        "Data", "Data de término",
        "Total por colaborador", "Controle de portaria"
    ]

    df = pd.read_excel(file)

    for column in COLUMNS:
        if column not in df.columns:
            raise KeyError(f"A coluna '{column}' não foi encontrada na planilha.")

    return df

def limpeza_df(df):
    df.columns = df.columns.str.strip()

    df["Colaborador"] = df["Colaborador"].fillna(method="ffill").astype(str).str.strip()
    df["Executores"] = df["Executores"].fillna("").astype(str).str.strip()
    df["Modelo de checklist"] = df["Modelo de checklist"].astype(str).str.strip()
    df["Controle de portaria"] = df["Controle de portaria"].fillna("").astype(str).str.strip()

    return df

def verificar_e_consolidar_divergencias(df):
    df["Erro de Controle de Portaria"] = df["Controle de portaria"].isna() | \
                                        (df["Controle de portaria"] == "-") | \
                                        (df["Controle de portaria"] == "")

    def contar_erros_execucao(grupo):
        modelos = grupo["Modelo de checklist"].tolist()
        erros = sum(1 for i in range(1, len(modelos)) if modelos[i] == modelos[i - 1])

        return pd.Series({"Erro na execução do Checklist": erros})

    erros_execucao = df.groupby("Executores").apply(contar_erros_execucao).reset_index()

    erros_executor = df[df["Executores"] != df["Colaborador"]].copy()
    erros_executor_counts = erros_executor.groupby("Executores").size().reset_index(name="Erro de Executor diferente de colaborador")

    erros_portaria_executor = df.groupby("Executores")["Erro de Controle de Portaria"].sum().reset_index()

    consolidado_colaborador = erros_execucao \
        .merge(erros_portaria_executor, on="Executores", how="left") \
        .merge(erros_executor_counts, on="Executores", how="left")

    consolidado_colaborador.drop(index=0, inplace=True)

    consolidado_colaborador.fillna(0, inplace=True)

    for column in ["Erro na execução do Checklist", "Erro de Executor diferente de colaborador", "Erro de Controle de Portaria"]:
        consolidado_colaborador[column] = consolidado_colaborador[column].astype(int)

    consolidado_colaborador["Total de Erros Colaborador"] = consolidado_colaborador[[
        "Erro na execução do Checklist",
        "Erro de Executor diferente de colaborador",
        "Erro de Controle de Portaria"
    ]].sum(axis=1)

    consolidado_colaborador.sort_values(by="Total de Erros Colaborador", ascending=False, inplace=True)

    total_erros_checklist = consolidado_colaborador["Erro na execução do Checklist"].sum()
    total_erros_executor = consolidado_colaborador["Erro de Executor diferente de colaborador"].sum()
    total_erros_portaria = consolidado_colaborador["Erro de Controle de Portaria"].sum()
    total_geral_erros = consolidado_colaborador["Total de Erros Colaborador"].sum()

    records_colaborador = []
    records_colaborador.append({
        "Rótulos de Linha": "TOTAL GERAL DE ERROS",
        "Contagem de Tipo de Erro": total_geral_erros
    })
    records_colaborador.append({
        "Rótulos de Linha": "  TOTAL ERROS NA EXECUÇÃO DO CHECKLIST",
        "Contagem de Tipo de Erro": total_erros_checklist
    })
    records_colaborador.append({
        "Rótulos de Linha": "  TOTAL ERROS DE EXECUTOR DIFERENTE",
        "Contagem de Tipo de Erro": total_erros_executor
    })
    records_colaborador.append({
        "Rótulos de Linha": "  TOTAL ERROS DE CONTROLE DE PORTARIA",
        "Contagem de Tipo de Erro": total_erros_portaria
    })

    for i, row in consolidado_colaborador.iterrows():
        records_colaborador.append({
            "Rótulos de Linha": row["Executores"],
            "Contagem de Tipo de Erro": row["Total de Erros Colaborador"]
        })
        records_colaborador.append({
            "Rótulos de Linha": "  Erro na execução do Checklist",
            "Contagem de Tipo de Erro": row["Erro na execução do Checklist"]
        })
        records_colaborador.append({
            "Rótulos de Linha": "  Erro de Executor diferente de colaborador",
            "Contagem de Tipo de Erro": row["Erro de Executor diferente de colaborador"]
        })
        records_colaborador.append({
            "Rótulos de Linha": "  Erros de Controle de Portaria",
            "Contagem de Tipo de Erro": row["Erro de Controle de Portaria"]
        })

    df_final_colaborador = pd.DataFrame(records_colaborador)

    return df_final_colaborador

def criar_planilha_excel(df):
    wb = Workbook()
    ws = wb.active
    ws.title = "Divergências por Colaborador"

    cabecalho_fill = PatternFill(start_color="548235", end_color="548235", fill_type="solid")
    cabecalho_font = Font(bold=True, color="FFFFFF")
    cabecalho_alignment = Alignment(horizontal="center", vertical="center")

    colaborador_fill = PatternFill(start_color="A9D18E", end_color="A9D18E", fill_type="solid")
    colaborador_font = Font(bold=True, color="FFFFFF")

    erro_fill = PatternFill(start_color="FFFFFF", end_color="FFFFFF", fill_type="solid")

    for col_num, value in enumerate(df.columns, start=1):
        cell = ws.cell(row=1, column=col_num, value=value)
        cell.fill = cabecalho_fill
        cell.font = cabecalho_font
        cell.alignment = cabecalho_alignment

    for row_num, row in enumerate(dataframe_to_rows(df, index=False, header=False), start=2):
        for col_num, cell_value in enumerate(row, start=1):
            cell = ws.cell(row=row_num, column=col_num, value=cell_value)

            if pd.notnull(row[0]) and not str(row[0]).startswith("  "):
                cell.fill = colaborador_fill
                cell.font = colaborador_font
            else:
                cell.fill = erro_fill

    col_widths = {
        1: 41,
        2: 28
    }

    for col_num, width in col_widths.items():
        col_letter = ws.cell(row=1, column=col_num).column_letter
        ws.column_dimensions[col_letter].width = width

    return wb

st.title("Análise de check-list")

def exportar_para_excel(df_final_colaborador, OUTPUT_PATH):
    with pd.ExcelWriter(OUTPUT_PATH, engine='xlsxwriter') as writer:
        df_final_colaborador.to_excel(writer, sheet_name='Consolidado_Colaborador', index=False)

def criar_excel_data(df_final_colaborador):
    excel_data = BytesIO()
    with pd.ExcelWriter(excel_data, engine='xlsxwriter') as writer:
        df_final_colaborador.to_excel(writer, sheet_name='Consolidado_Colaborador', index=False)
    excel_data.seek(0)
    return excel_data

uploaded_file = st.file_uploader("Escolha o arquivo Excel", type="xlsx")

if uploaded_file is not None:
    df = importar_planilha(uploaded_file)
    df = limpeza_df(df)
    df_final_colaborador = verificar_e_consolidar_divergencias(df)

# Definir as funções de plotagem após criar_excel_data
def plotar_grafico(df_final_colaborador):
    grafico_data = df_final_colaborador.iloc[:4].copy()
    grafico_data.set_index("Rótulos de Linha", inplace=True)

    fig, ax = plt.subplots(figsize=(8, 4))
    grafico_data["Contagem de Tipo de Erro"].plot(kind='bar', ax=ax)
    ax.set_title("Total de Erros e Tipos de Erros")
    ax.set_ylabel("Contagem de Erros")
    ax.set_xlabel("Tipos de Erros")

    return fig

def plotar_grafico_adicional(df_final_colaborador):
    indices = [4, 8, 12, 16, 20, 24, 28, 32, 36, 40, 44, 48, 52, 56, 60]
    grafico_adicional_data = df_final_colaborador.iloc[indices].copy()
    grafico_adicional_data.set_index("Rótulos de Linha", inplace=True)

    fig, ax = plt.subplots(figsize=(8, 4))
    grafico_adicional_data["Contagem de Tipo de Erro"].plot(kind='bar', ax=ax, color='skyblue', edgecolor='black')
    ax.set_title("Comparativo 15 colaboradores com mais erros")
    ax.set_ylabel("Contagem de Erros")
    ax.set_xlabel("Colaboradores")
    ax.set_xticklabels(ax.get_xticklabels(), rotation=45, ha="right")
    ax.grid(True, which='both', linestyle='--', linewidth=0.5)
    ax.legend(["Total de Erros"], loc='upper right')

    return fig

# Dentro do if uploaded_file is not None:
if uploaded_file is not None:
    df = importar_planilha(uploaded_file)
    df = limpeza_df(df)
    df_final_colaborador = verificar_e_consolidar_divergencias(df)

    st.success("Relatório de divergências gerado com sucesso!")
    
    # Adicionar a pré-visualização do DataFrame corrigido 
    st.dataframe(df_final_colaborador)

    # Chamada das funções de plotagem
    fig1 = plotar_grafico(df_final_colaborador)
    fig2 = plotar_grafico_adicional(df_final_colaborador)
    
    # Exibir os gráficos
    st.pyplot(fig1)
    st.pyplot(fig2)

    excel_data = criar_excel_data(df_final_colaborador)
    
    st.download_button(
        label="Download do arquivo corrigido",
        data=excel_data,
        file_name="relatorio_divergencias_corrigido.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
